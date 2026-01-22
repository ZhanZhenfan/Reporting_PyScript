import os
import sys
from datetime import datetime, timedelta, date

INVENTORY_DIR = r"\\mygbynbyn1msis1\Supply-Chain-Analytics\Data Warehouse\Data Source\SAP\Transactional Data\Inventory"

# 只处理这6个文件名（区分 .xlsx / .xls 自动适配）
TARGET_BASENAMES = {
    "CN MB52", "KKAQ", "MB5T", "MY MB52", "SG MB52", "US MB52"
}

def this_monday() -> date:
    today = datetime.now().date()
    return today - timedelta(days=today.weekday())  # Monday=0

# ---------- XLSX 处理（openpyxl） ----------
def process_xlsx(path: str):
    from openpyxl import load_workbook

    wb = load_workbook(path)
    ws = wb.active

    # 找最后一列（openpyxl 的 max_column 就是已用的最右列）
    last_col = ws.max_column
    last_row = ws.max_row

    # 估计是否首行是表头：如果首行该列是字符串就认为是表头，从第2行开始
    header_cell = ws.cell(row=1, column=last_col)
    start_row = 2 if isinstance(header_cell.value, str) else 1

    # 读取该列原本的日期格式（找第一个非空数据单元格）
    num_fmt = None
    probe_end = min(start_row + 100, last_row + 1)
    for r in range(start_row, probe_end):
        c = ws.cell(row=r, column=last_col)
        if c.value not in (None, ""):
            # 如果这一列原来就是日期/数字，直接沿用它的 number_format
            fmt = getattr(c, "number_format", None)
            if fmt and fmt != "General":
                num_fmt = fmt
            break
    if not num_fmt:
        # 默认一个安全的通用格式（Excel 会按系统区域展示）
        num_fmt = "yyyy-mm-dd"

    # 写入整列为本周一，并设置格式
    monday = this_monday()
    for r in range(start_row, last_row + 1):
        # 简单防脏：如果整行几乎是空的，可以跳过；否则统一写
        row_has_data = any(ws.cell(row=r, column=c).value not in (None, "") for c in range(1, max(2, last_col)))
        if not row_has_data:
            continue
        cell = ws.cell(row=r, column=last_col)
        cell.value = monday
        cell.number_format = num_fmt

    wb.save(path)
    print(f"✓ XLSX done: {os.path.basename(path)}  (last col #{last_col}, rows {start_row}-{last_row}, fmt='{num_fmt}')")

# ---------- XLS 处理（Excel COM） ----------
def process_xls(path: str):
    try:
        import win32com.client as win32
    except ImportError:
        print(f"! 跳过（缺少 pywin32）：{path} / Skipped (missing pywin32): {path}")
        return

    excel = win32.gencache.EnsureDispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False

    try:
        wb = excel.Workbooks.Open(path)
        ws = wb.ActiveSheet

        used = ws.UsedRange
        last_row = used.Rows.Count
        last_col = used.Columns.Count

        # 估计是否表头
        header_val = ws.Cells(1, last_col).Value
        start_row = 2 if isinstance(header_val, str) else 1

        # 获取列格式（优先从首个有值的单元格读）
        num_fmt = None
        for r in range(start_row, min(start_row + 100, last_row + 1)):
            v = ws.Cells(r, last_col).Value
            if v is not None and v != "":
                fmt = ws.Cells(r, last_col).NumberFormat
                if fmt and fmt != "General":
                    num_fmt = fmt
                break
        if not num_fmt:
            num_fmt = "yyyy-mm-dd"

        monday = this_monday()
        # 写整列
        for r in range(start_row, last_row + 1):
            # 行是否有数据
            row_has_data = False
            for c in range(1, max(2, last_col)):
                if ws.Cells(r, c).Value not in (None, ""):
                    row_has_data = True
                    break
            if not row_has_data:
                continue

            ws.Cells(r, last_col).Value = monday
            ws.Cells(r, last_col).NumberFormat = num_fmt

        wb.Save()
        print(f"✓ XLS  done: {os.path.basename(path)}  (last col #{last_col}, rows {start_row}-{last_row}, fmt='{num_fmt}')")
    finally:
        wb.Close(SaveChanges=True)
        excel.Quit()

def main():
    if not os.path.isdir(INVENTORY_DIR):
        print("目录不存在：", INVENTORY_DIR, "/ Directory not found:", INVENTORY_DIR)
        sys.exit(1)

    # 根据文件名（不含扩展名）匹配这 6 个目标
    files = os.listdir(INVENTORY_DIR)
    wanted = []
    for f in files:
        base, ext = os.path.splitext(f)
        if ext.lower() not in (".xlsx", ".xls"):
            continue
        if base in TARGET_BASENAMES:
            wanted.append(os.path.join(INVENTORY_DIR, f))

    # 如果目录里文件名带空格差异（如 “CN MB52.xlsx” / “CN MB52 .xlsx”），也可放宽匹配：
    if not wanted:
        for f in files:
            base, ext = os.path.splitext(f)
            if ext.lower() not in (".xlsx", ".xls"):
                continue
            tight = base.replace(" ", "")
            for name in TARGET_BASENAMES:
                if tight == name.replace(" ", ""):
                    wanted.append(os.path.join(INVENTORY_DIR, f))
                    break

    if not wanted:
        print("未找到目标文件。请确认目录下存在以下任一文件名（.xlsx/.xls）："
              " / No target files found. Please verify one of these filenames exists (.xlsx/.xls):")
        for n in sorted(TARGET_BASENAMES):
            print(" -", n)
        sys.exit(1)

    for path in sorted(wanted):
        ext = os.path.splitext(path)[1].lower()
        try:
            if ext == ".xlsx":
                process_xlsx(path)
            else:
                process_xls(path)
        except Exception as e:
            print(f"! 处理失败：{os.path.basename(path)} -> {e} / Failed to process: {os.path.basename(path)} -> {e}")

    print("✅ 完成：已将最后一列改为本周周一，并沿用原列日期格式。 / "
          "Done: last column set to this Monday with original date format.")

if __name__ == "__main__":
    main()
