import os, re, time, shutil
from typing import Optional, List
from pathlib import Path

import pandas as pd  # 仅占位；当前逻辑未直接用到
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# ============ 引入你的两个工具类 ============
from Utils.graph_mail_attachment_tool import GraphMailAttachmentTool
from Utils.sql_agent_tool import SqlAgentTool

# ============
# 配置区（按需修改）
# ============

# Graph 应用信息
TENANT_ID = "5c2be51b-4109-461d-a0e7-521be6237ce2"
CLIENT_ID = "09004044-1c60-48e5-b1eb-bb42b3892006"

# 邮件附件匹配规则
ATTACHMENT_NAME_EQUALS: Optional[str] = None
ATTACHMENT_NAME_CONTAINS = "ZMRP_WATERFALL_Run"
ATTACHMENT_EXT = ".xlsx"

# 下载行为
NEED_COUNT = 1
DAYS_BACK  = 90
PAGE_SIZE  = 50
MAX_SCAN   = 800
MAIL_FOLDER = "inbox"   # 不限制可设为 None

# 本地目录
LOCAL_TMP_DIR   = r"\\mp1do4ce0373ndz\C\WeeklyRawFile\Download_From_Eamil"
LOCAL_CLEAN_DIR = r"\\mp1do4ce0373ndz\C\WeeklyRawFile\Download_From_Eamil\Processed"

# 共享盘目录与最终文件名
SHARE_DEST_DIR = r"\\mygbynbyn1msis1\Supply-Chain-Analytics\Data Warehouse\Data Source\SAP\Transactional Data\MRP Waterfall"
DEST_FILENAME  = "Month MY0X ZMRP_WATERFALL.xlsx"

# “占位/中间文件”关键字与等待参数
BLOCKING_NAME_KEYWORDS = ["W#1"]
WAIT_TIMEOUT_SEC = 45 * 60
WAIT_POLL_SEC    = 10

# ---- SQL Job（留空给你填） ----
SQL_SERVER   = "10.80.127.71,1433"   # 例： "tcp:10.80.127.71,1433"
SQL_JOB_NAME = "Lumileds BI - SC RawMaterialEOHProjection"   # 例： "Lumileds BI - SC MRP Waterfall"（建议精确名）
ARCHIVE_DIR = r"\\mygbynbyn1msis1\Supply-Chain-Analytics\Data Warehouse\Data Source\SAP\Transactional Data\MRP Waterfall"


# ============ 小工具 ============

def ensure_dir(p: str):
    Path(p).mkdir(parents=True, exist_ok=True)

def newest_file(paths: List[str]) -> Optional[str]:
    files = [p for p in paths if p and os.path.isfile(p)]
    return max(files, key=lambda p: os.path.getmtime(p)) if files else None

def wait_folder_clear(folder: str, keywords: List[str], timeout_sec: int, poll_sec: int) -> bool:
    print(f"⏳ 等待共享盘清空占位文件（关键词：{keywords}）...")
    t0 = time.time()
    while True:
        try:
            names = os.listdir(folder)
        except FileNotFoundError:
            names = []
        blocked = [n for n in names for k in keywords if k.lower() in n.lower()]
        if not blocked:
            print("✅ 共享盘状态良好，可复制。")
            return True
        if time.time() - t0 > timeout_sec:
            print(f"⚠ 超时仍存在：{blocked[:5]} ...")
            return False
        time.sleep(poll_sec)

def _normalize_material_text(s: str) -> str:
    """
    物料号文本规范化：
      - 去首尾空格
      - '12345.0' -> '12345'
      - 仅当是数字且以 '00000' 开头时，移除**前 5 个 0**
    """
    s = "" if s is None else str(s).strip()
    if not s:
        return ""
    import re
    if re.fullmatch(r"\d+(\.0+)?", s):
        try:
            s = str(int(float(s)))
        except Exception:
            pass
    if s.startswith("00000") and s[5:].isdigit():
        s = s[5:]
    return s

def clean_workbook(in_xlsx: str, out_xlsx: str):
    """
    用 openpyxl 清洗：
      - A列：文本格式，按 _normalize_material_text 规范化
      - E列：尽量转为数值（其余保留）
    """
    print(f"🧽 清洗（保物料号）：{in_xlsx}")
    wb = load_workbook(in_xlsx, data_only=True)
    ws = wb.active

    max_row = ws.max_row

    # A 列：文本
    for r in range(1, max_row + 1):
        c = ws.cell(row=r, column=1)
        c.value = _normalize_material_text(c.value)
        c.number_format = '@'

    # E 列：数值
    for r in range(1, max_row + 1):
        c = ws.cell(row=r, column=5)
        val = c.value
        if val in (None, ""):
            continue
        sval = str(val).strip().replace(",", "")
        try:
            f = float(sval)
            c.value = int(f) if f.is_integer() else f
            c.number_format = "0"
        except Exception:
            pass

    ensure_dir(os.path.dirname(out_xlsx) or ".")
    wb.save(out_xlsx)
    print(f"✔ 清洗完成 -> {out_xlsx}")

def copy_to_share(src_file: str, dest_folder: str) -> str:
    """
    另存为固定文件名并复制到共享盘（直接覆盖，不做备份）。
    """
    dest_path = os.path.join(dest_folder, DEST_FILENAME)
    # 确保目录存在（一般共享盘已存在，这里稳妥一下）
    Path(dest_folder).mkdir(parents=True, exist_ok=True)

    # 直接覆盖（shutil.copy2 遇到同名文件会覆盖）
    shutil.copy2(src_file, dest_path)
    print(f"📤 已复制并覆盖共享盘：{dest_path}")
    return dest_path


# ============ 主流程：下载 + 清洗 + 复制 + 触发Job ============

def main():
    # Step 1：下载（调用工具类）
    print("==== Step 1: 从邮箱下载月度文件 ====")
    ensure_dir(LOCAL_TMP_DIR)
    graph_tool = GraphMailAttachmentTool(
        tenant_id=TENANT_ID,
        client_id=CLIENT_ID
    )
    saved_paths = graph_tool.download_latest_attachments(
        contains=ATTACHMENT_NAME_CONTAINS if not ATTACHMENT_NAME_EQUALS else None,
        equals=ATTACHMENT_NAME_EQUALS,
        ext=ATTACHMENT_EXT,
        need_count=NEED_COUNT,
        days_back=DAYS_BACK,
        page_size=PAGE_SIZE,
        max_scan=MAX_SCAN,
        save_dir=LOCAL_TMP_DIR,
        mail_folder=MAIL_FOLDER,
    )
    latest_raw = newest_file([str(p) for p in saved_paths])
    if not latest_raw:
        raise RuntimeError("未获取到任何附件文件。")
    print(f"➡ 最新原始文件：{latest_raw}")

    # Step 2 & 3：清洗
    print("\n==== Step 2 & 3: 另存并清洗 ====")
    ensure_dir(LOCAL_CLEAN_DIR)
    cleaned_tmp = os.path.join(
        LOCAL_CLEAN_DIR,
        os.path.splitext(os.path.basename(latest_raw))[0] + ".cleaned.xlsx"
    )
    clean_workbook(latest_raw, cleaned_tmp)

    # Step 4：等待共享盘空闲并复制
    print("\n==== Step 4: 复制到共享盘（含占位检查） ====")
    ok = wait_folder_clear(SHARE_DEST_DIR, BLOCKING_NAME_KEYWORDS, WAIT_TIMEOUT_SEC, WAIT_POLL_SEC)
    if not ok:
        print("⚠ 未能确认共享盘空闲。为安全起见，本次不复制。你可以稍后手动把下列文件放进去：")
        print(f"   {cleaned_tmp}")
        return
    dest = copy_to_share(cleaned_tmp, SHARE_DEST_DIR)

    # Step 5：触发 SQL Job（用工具类；会哔哔声并打开 Archive 文件夹）
    print("\n==== Step 5: 触发 SQL Job ====")
    if SQL_SERVER and SQL_JOB_NAME and ARCHIVE_DIR:
        sql_tool = SqlAgentTool(server=SQL_SERVER)
        result = sql_tool.run_job(
            job_name=SQL_JOB_NAME,
            archive_dir=ARCHIVE_DIR,   # 成功后会打开此文件夹
            timeout=1800,
            poll_interval=3,
            fuzzy=False,               # 若之后有读 sysjobs 权限，可设 True
        )
        print("[JOB RESULT]", result)
    else:
        print("（跳过 Job：请在配置区填写 SQL_SERVER / SQL_JOB_NAME / ARCHIVE_DIR 后启用）")

    print("\n✅ 全流程完成。")
    print("原始下载：", latest_raw)
    print("清洗临时：", cleaned_tmp)
    print("共享盘路径：", dest)

if __name__ == "__main__":
    main()
