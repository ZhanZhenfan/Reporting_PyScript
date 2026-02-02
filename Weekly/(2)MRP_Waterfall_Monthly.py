# -*- coding: utf-8 -*-
import os, re, time, shutil, glob
from typing import Optional, List, Tuple
from pathlib import Path

import pandas as pd  # 仅占位
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# ============ 引入你的两个工具类 ============
from Utils.graph_mail_attachment_tool import GraphMailAttachmentTool
from Utils.sql_agent_tool import SqlAgentTool
from Utils.email_notify_tool import EmailNotifier
# ===========================================

# ----------------- 全局开关：输入源 -----------------
# "email" -> 从邮箱下载到 LOCAL_TMP_DIR，再用最新文件
# "folder" -> 直接在 FOLDER_SOURCE_DIR 中按匹配规则挑最新文件
INPUT_MODE: str = os.getenv("MRP_INPUT_MODE", "email").lower()  # email | folder
# ---------------------------------------------------

# ============ 配置区（按需修改） ============

# Graph 应用信息（仅在 INPUT_MODE='email' 有效）
TENANT_ID = "5c2be51b-4109-461d-a0e7-521be6237ce2"
CLIENT_ID = "09004044-1c60-48e5-b1eb-bb42b3892006"

# 附件/文件匹配规则（两种模式共用）
ATTACHMENT_NAME_EQUALS: Optional[str] = None           # 精确名优先
ATTACHMENT_NAME_CONTAINS = "ZMRP_WATERFALL_M"          # 其次用“包含”匹配
ATTACHMENT_EXT = ".xlsx"                               # 扩展名过滤

# 下载行为（仅在 INPUT_MODE='email' 有效）
NEED_COUNT = 1
DAYS_BACK  = 90
PAGE_SIZE  = 50
MAX_SCAN   = 800
MAIL_FOLDER = "inbox"   # 不限制可设为 None

# 本地目录（两种模式都会用到）
LOCAL_TMP_DIR   = r"\\mp1do4ce0373ndz\C\WeeklyRawFile\Download_From_Eamil"
LOCAL_CLEAN_DIR = r"\\mp1do4ce0373ndz\C\WeeklyRawFile\Download_From_Eamil\Processed"

# 当 INPUT_MODE='folder' 时，从这个目录直接找原始文件
FOLDER_SOURCE_DIR = LOCAL_TMP_DIR  # 也可单独设一个目录
# 可用 glob 模式做进一步过滤（为空则用 equals/contains/ext 组合规则）
FOLDER_GLOB_PATTERNS: List[str] = []  # 例如：["*ZMRP_WATERFALL*.xlsx"]

# 共享盘目录与最终文件名
SHARE_DEST_DIR = r"\\mygbynbyn1msis1\Supply-Chain-Analytics\Data Warehouse\Data Source\SAP\Transactional Data\MRP Waterfall"
DEST_FILENAME  = "Month MY0X ZMRP_WATERFALL.xlsx"

# “占位/中间文件”关键字与等待参数
BLOCKING_NAME_KEYWORDS = ["W#1"]
WAIT_TIMEOUT_SEC = 45 * 60
WAIT_POLL_SEC    = 10

# ---- SQL Job ----
SQL_SERVER   = "10.80.127.71,1433"
SQL_JOB_NAME = "Lumileds BI - SC RawMaterialEOHProjection"
ARCHIVE_DIR  = r"\\mygbynbyn1msis1\Supply-Chain-Analytics\Data Warehouse\Data Source\SAP\Transactional Data\MRP Waterfall\Archive"

# ---- Email notify (optional) ----
ENABLE_EMAIL_NOTIFY = os.getenv("EMAIL_NOTIFY", "0").strip().lower() in {"1", "true", "yes"}
JOB_KEY = "MRP_Waterfall_Monthly"

# Job-specific message templates (customize per job)
SUCCESS_SUBJECT = "MRP Waterfall Monthly - Success"
SUCCESS_BODY = "MRP Waterfall Monthly completed successfully."
FAIL_SUBJECT = "MRP Waterfall Monthly - Failed"
FAIL_BODY_PREFIX = "MRP Waterfall Monthly failed with error:\n"


# ============ 小工具 ============

def ensure_dir(p: str):
    Path(p).mkdir(parents=True, exist_ok=True)

def newest_file(paths: List[str]) -> Optional[str]:
    files = [p for p in paths if p and os.path.isfile(p)]
    return max(files, key=lambda p: os.path.getmtime(p)) if files else None


def _notify(subject: str, body: str) -> None:
    if not ENABLE_EMAIL_NOTIFY:
        return
    notifier = EmailNotifier.from_config()
    notifier.send_with_config(
        job_key=JOB_KEY,
        subject=subject,
        body=body,
    )

def list_matching_files_in_dir(
    folder: str,
    equals: Optional[str],
    contains: Optional[str],
    ext: Optional[str],
    extra_globs: Optional[List[str]] = None
) -> List[str]:
    """
    在 folder 中返回满足条件的文件列表（不递归）。
    优先顺序只是匹配选择逻辑，不做排序；排序交给 newest_file。
    """
    try:
        names = [n for n in os.listdir(folder) if os.path.isfile(os.path.join(folder, n))]
    except FileNotFoundError:
        return []

    candidates: List[str] = []

    # 如果配置了额外 glob，则直接按 glob 拿（允许多模式）
    if extra_globs:
        for pat in extra_globs:
            candidates.extend(glob.glob(os.path.join(folder, pat)))
        # 去重
        candidates = list({os.path.abspath(p) for p in candidates if os.path.isfile(p)})
        return candidates

    # 否则用 equals / contains / ext 的规则
    # 1) equals（精确名）
    if equals:
        for n in names:
            if n == equals:
                candidates.append(os.path.join(folder, n))
        if candidates:
            return candidates

    # 2) contains + ext
    for n in names:
        ok_contains = (contains.lower() in n.lower()) if contains else True
        ok_ext = n.lower().endswith(ext.lower()) if ext else True
        if ok_contains and ok_ext:
            candidates.append(os.path.join(folder, n))

    return candidates

def wait_folder_clear(folder: str, keywords: List[str], timeout_sec: int, poll_sec: int) -> bool:
    print(f"⏳ 等待共享盘清空占位文件（关键词：{keywords}）... / Waiting for share to clear blocking files (keywords: {keywords})...")
    t0 = time.time()
    while True:
        try:
            names = os.listdir(folder)
        except FileNotFoundError:
            names = []
        blocked = [n for n in names for k in keywords if k.lower() in n.lower()]
        if not blocked:
            print("✅ 共享盘状态良好，可复制。 / Share is clear; ready to copy.")
            return True
        if time.time() - t0 > timeout_sec:
            print(f"⚠ 超时仍存在：{blocked[:5]} ... / Timeout; still blocked: {blocked[:5]} ...")
            return False
        time.sleep(poll_sec)

def _normalize_material_text(s: str) -> str:
    s = "" if s is None else str(s).strip()
    if not s:
        return ""
    if re.fullmatch(r"\d+(\.0+)?", s):
        try:
            s = str(int(float(s)))
        except Exception:
            pass
    if s.startswith("00000") and s[5:].isdigit():
        s = s[5:]
    return s

def clean_workbook(in_xlsx: str, out_xlsx: str):
    print(f"🧽 清洗（保物料号）：{in_xlsx} / Cleaning (keep material number): {in_xlsx}")
    wb = load_workbook(in_xlsx, data_only=True)
    ws = wb.active

    max_row = ws.max_row

    # A 列：文本
    for r in range(1, max_row + 1):
        c = ws.cell(row=r, column=1)
        c.value = _normalize_material_text(c.value)
        c.number_format = '@'

    # E 列：数值
    for r in range(1, max_row + 1):
        c = ws.cell(row=r, column=5)
        val = c.value
        if val in (None, ""):
            continue
        sval = str(val).strip().replace(",", "")
        try:
            f = float(sval)
            c.value = int(f) if f.is_integer() else f
            c.number_format = "0"
        except Exception:
            pass

    ensure_dir(os.path.dirname(out_xlsx) or ".")
    wb.save(out_xlsx)
    print(f"✔ 清洗完成 -> {out_xlsx} / Cleaning done -> {out_xlsx}")

def copy_to_share(src_file: str, dest_folder: str) -> str:
    dest_path = os.path.join(dest_folder, DEST_FILENAME)
    Path(dest_folder).mkdir(parents=True, exist_ok=True)
    shutil.copy2(src_file, dest_path)
    print(f"📤 已复制并覆盖共享盘：{dest_path} / Copied and replaced on share: {dest_path}")
    return dest_path


# ============ 输入源解耦 ============

def fetch_from_email() -> str:
    """
    从邮箱下载到 LOCAL_TMP_DIR，返回最新文件路径。
    """
    print("==== Step 1: 从邮箱下载月度文件 / Download monthly file from email ====")
    ensure_dir(LOCAL_TMP_DIR)
    graph_tool = GraphMailAttachmentTool(
        tenant_id=TENANT_ID,
        client_id=CLIENT_ID
    )
    saved_paths = graph_tool.download_latest_attachments(
        contains=ATTACHMENT_NAME_CONTAINS if not ATTACHMENT_NAME_EQUALS else None,
        equals=ATTACHMENT_NAME_EQUALS,
        ext=ATTACHMENT_EXT,
        need_count=NEED_COUNT,
        days_back=DAYS_BACK,
        page_size=PAGE_SIZE,
        max_scan=MAX_SCAN,
        save_dir=LOCAL_TMP_DIR,
        mail_folder=MAIL_FOLDER,
    )

    # saved_paths 已经是下载得到的文件；兜底再在目录中按规则找一遍
    candidates = []
    if saved_paths:
        candidates.extend([str(p) for p in saved_paths if p and os.path.isfile(str(p))])

    if not candidates:
        candidates = list_matching_files_in_dir(
            folder=LOCAL_TMP_DIR,
            equals=ATTACHMENT_NAME_EQUALS,
            contains=ATTACHMENT_NAME_CONTAINS,
            ext=ATTACHMENT_EXT,
            extra_globs=None
        )

    latest = newest_file(candidates)
    if not latest:
        raise RuntimeError("未获取到任何附件文件。 / No attachments were retrieved.")
    print(f"➡ 最新原始文件（邮箱）：{latest} / Latest raw file (email): {latest}")
    return latest

def fetch_from_folder() -> str:
    """
    直接在 FOLDER_SOURCE_DIR 中找匹配的最新文件，返回路径。
    """
    print("==== Step 1: 从文件夹选择最新文件 / Pick latest file from folder ====")
    ensure_dir(FOLDER_SOURCE_DIR)
    candidates = list_matching_files_in_dir(
        folder=FOLDER_SOURCE_DIR,
        equals=ATTACHMENT_NAME_EQUALS,
        contains=ATTACHMENT_NAME_CONTAINS,
        ext=ATTACHMENT_EXT,
        extra_globs=FOLDER_GLOB_PATTERNS or None
    )
    latest = newest_file(candidates)
    if not latest:
        hint = f"目录为空或无匹配：{FOLDER_SOURCE_DIR} / Folder empty or no match: {FOLDER_SOURCE_DIR}"
        if FOLDER_GLOB_PATTERNS:
            hint += f"；glob={FOLDER_GLOB_PATTERNS} / glob={FOLDER_GLOB_PATTERNS}"
        else:
            hint += (
                f"；规则=equals:{ATTACHMENT_NAME_EQUALS} / contains:{ATTACHMENT_NAME_CONTAINS} / ext:{ATTACHMENT_EXT}"
                f" / rules=equals:{ATTACHMENT_NAME_EQUALS} / contains:{ATTACHMENT_NAME_CONTAINS} / ext:{ATTACHMENT_EXT}"
            )
        raise RuntimeError(hint)
    print(f"➡ 最新原始文件（文件夹）：{latest} / Latest raw file (folder): {latest}")
    return latest

def get_latest_input() -> str:
    """
    根据 INPUT_MODE 选择输入源，并返回“原始文件路径”。
    """
    mode = INPUT_MODE
    if mode not in ("email", "folder"):
        print(f"⚠ 未知 INPUT_MODE={mode}，回退到 'folder' / Unknown INPUT_MODE={mode}, falling back to 'folder'")
        mode = "folder"

    if mode == "email":
        return fetch_from_email()
    else:
        return fetch_from_folder()


# ============ 主流程：清洗 + 复制 + 触发Job ============

def main():
    print(f"==== MRP Waterfall（输入源：{INPUT_MODE}）==== / MRP Waterfall (source: {INPUT_MODE}) ====")

    # Step 1：拿到“原始文件”
    latest_raw = get_latest_input()

    # Step 2 & 3：清洗
    print("\n==== Step 2 & 3: 另存并清洗 / Save as and clean ====")
    ensure_dir(LOCAL_CLEAN_DIR)
    cleaned_tmp = os.path.join(
        LOCAL_CLEAN_DIR,
        os.path.splitext(os.path.basename(latest_raw))[0] + ".cleaned.xlsx"
    )
    clean_workbook(latest_raw, cleaned_tmp)

    # Step 4：等待共享盘空闲并复制
    print("\n==== Step 4: 复制到共享盘（含占位检查） / Copy to share (with blocking check) ====")
    ok = wait_folder_clear(SHARE_DEST_DIR, BLOCKING_NAME_KEYWORDS, WAIT_TIMEOUT_SEC, WAIT_POLL_SEC)
    if not ok:
        print("⚠ 未能确认共享盘空闲。为安全起见，本次不复制。你可以稍后手动把下列文件放进去："
              " / Share not confirmed clear; skip copy for safety. You can manually place this file later:")
        print(f"   {cleaned_tmp}")
        return
    dest = copy_to_share(cleaned_tmp, SHARE_DEST_DIR)

    # Step 5：触发 SQL Job
    print("\n==== Step 5: 触发 SQL Job / Trigger SQL Job ====")
    if SQL_SERVER and SQL_JOB_NAME and ARCHIVE_DIR:
        sql_tool = SqlAgentTool(server=SQL_SERVER)
        result = sql_tool.run_job(
            job_name=SQL_JOB_NAME,
            archive_dir=ARCHIVE_DIR,
            timeout=1800,
            poll_interval=3,
            fuzzy=False,
        )
        print("[JOB RESULT]", result)
    else:
        print("（跳过 Job：请在配置区填写 SQL_SERVER / SQL_JOB_NAME / ARCHIVE_DIR 后启用） / "
              "Job skipped: fill SQL_SERVER / SQL_JOB_NAME / ARCHIVE_DIR in config to enable.")

    print("\n✅ 全流程完成。 / Full workflow completed.")
    print("原始下载：", latest_raw, "/ Raw download:", latest_raw)
    print("清洗临时：", cleaned_tmp, "/ Cleaned temp:", cleaned_tmp)
    print("共享盘路径：", dest, "/ Share path:", dest)

if __name__ == "__main__":
    try:
        main()
        _notify(SUCCESS_SUBJECT, SUCCESS_BODY)
    except BaseException as e:
        _notify(FAIL_SUBJECT, f"{FAIL_BODY_PREFIX}{e}")
        raise
