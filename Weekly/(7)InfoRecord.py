import os
import re
import time
import shutil
from datetime import datetime
from typing import Optional

from glob import glob
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from Utils.sql_agent_tool import SqlAgentTool

# ===== 配置 =====
SRC_DIR   = r"\\mygbynbyn1vw214\InfoRecord"
DEST_DIR  = r"\\mygbynbyn1msis1\Supply-Chain-Analytics\Data Warehouse\Data Source\SAP\Master Data"
DEST_NAME = "InfoRecord.xlsx"

SHEET_INDEX = 0           # 要处理的工作表（0 = 第一个）
COL_L_IDX   = 12          # L 列 = 第 12 列
INS_Q_IDX   = 17          # 在第 17 列（Q）插入空列
HEADER_Q    = "Crcy"      # 新列表头

# 目标目录“阻塞文件”关键词（存在则等待）
BLOCKING_KEYWORDS = ["Task6", "W#1"]
WAIT_TIMEOUT_SEC  = 20 * 60    # 最多等 20 分钟
WAIT_POLL_SEC     = 5          # 每 5 秒检查一次
# =================


def find_latest_inforecord(src_dir: str) -> Optional[str]:
    """
    从 src_dir 下选择“最新修改”的子文件夹，
    再在该子文件夹内选择最新的  PR1 Info Record*.xlsx  文件，返回完整路径。
    """
    if not os.path.isdir(src_dir):
        print(f"⚠ 路径不存在：{src_dir} / Path not found: {src_dir}")
        return None

    # 取所有子目录
    subdirs = [
        os.path.join(src_dir, d)
        for d in os.listdir(src_dir)
        if os.path.isdir(os.path.join(src_dir, d))
    ]
    if not subdirs:
        print("⚠ 未找到任何子文件夹。 / No subfolders found.")
        return None

    # 最新修改时间优先
    subdirs.sort(key=os.path.getmtime, reverse=True)
    newest_dir = subdirs[0]
    print(f"📁 最新子目录：{newest_dir} / Latest subfolder: {newest_dir}")

    # 仅匹配 “PR1 Info Record*.xlsx”，忽略 Excel 临时文件 "~$*.xlsx"
    cand_files = [
        os.path.join(newest_dir, f)
        for f in os.listdir(newest_dir)
        if f.lower().endswith(".xlsx")
        and f.lower().startswith("pr1 info record")
        and not f.startswith("~$")
    ]
    if not cand_files:
        print("⚠ 最新子目录内未找到 PR1 Info Record*.xlsx / No PR1 Info Record*.xlsx in latest subfolder")
        return None

    cand_files.sort(key=os.path.getmtime, reverse=True)
    latest_file = cand_files[0]
    print(f"✅ 选定源文件：{latest_file} / Selected source file: {latest_file}")
    return latest_file


def is_blocking_present(folder: str) -> list[str]:
    try:
        names = os.listdir(folder)
    except FileNotFoundError:
        return []
    hits = []
    for n in names:
        for k in BLOCKING_KEYWORDS:
            if k.lower() in n.lower():
                hits.append(n)
                break
    return hits


def wait_dest_clear(dest_dir: str) -> bool:
    print(f"⏳ 检查目标目录是否空闲：{dest_dir} / Checking if destination folder is clear: {dest_dir}")
    t0 = time.time()
    while True:
        hits = is_blocking_present(dest_dir)
        if not hits:
            print("  ✅ 目标目录无阻塞文件，可以写入。 / No blocking files; ready to write.")
            return True
        waited = int(time.time() - t0)
        print(f"  … 检测到阻塞文件：{hits[:5]}（已等 {waited}s） / "
              f"Blocking files detected: {hits[:5]} (waited {waited}s)")
        if time.time() - t0 > WAIT_TIMEOUT_SEC:
            print("  ⚠ 等待超时，仍存在阻塞文件。为安全起见，本次不覆盖落盘。 / "
                  "Timeout; blocking files remain. Skipping write for safety.")
            return False
        time.sleep(WAIT_POLL_SEC)


def remove_leading_zeros_keep_text(v):
    """
    仅对“全为数字”的内容去前导零；非数字、混合字符不动。
    返回字符串；空/None -> 空串
    """
    if v is None:
        return ""
    s = str(v).strip()
    if s == "":
        return ""
    # 去掉千分位逗号
    s_clean = s.replace(",", "")
    if re.fullmatch(r"\d+", s_clean):
        # 纯数字：转 int 再转回字符串 -> 去前导零
        return str(int(s_clean))
    # 其他：保持不变
    return s


def process_workbook_and_save(xlsx_path: str,
                              out_path: str,
                              sheet_index: int = SHEET_INDEX,
                              col_l_idx: int = COL_L_IDX,
                              ins_q_idx: int = INS_Q_IDX,
                              header_q: str = HEADER_Q) -> str:
    """
    直接把清洗后的内容保存到 out_path（不会在源目录生成任何临时文件）。
    """
    print(f"🔧 打开工作簿：{xlsx_path} / Opening workbook: {xlsx_path}")
    wb = load_workbook(xlsx_path)  # 不用 data_only，避免公式被提前求值
    ws = wb.worksheets[sheet_index]

    max_row = ws.max_row
    print(f"  工作表：{ws.title} | 行数≈{max_row} / Sheet: {ws.title} | Rows≈{max_row}")

    # Step 3：清洗 L 列（去前导0 + 设为文本）
    col_letter = get_column_letter(col_l_idx)
    print(f"  Step 3 | 处理列 {col_letter}：去前导 0（仅纯数字），并设置为文本格式 … / "
          f"Step 3 | Column {col_letter}: remove leading zeros (numeric only) and set text format ...")

    changed = 0
    for r in range(2, max_row + 1):
        cell = ws.cell(row=r, column=col_l_idx)
        new_val = remove_leading_zeros_keep_text(cell.value)
        if new_val != ("" if cell.value is None else str(cell.value).strip()):
            changed += 1
        cell.value = new_val
        cell.number_format = "@"  # 文本格式

    print(f"    ✔ L 列处理完成，改动约 {changed} 行。 / L column done, ~{changed} rows changed.")

    # Step 4：在 Q 列插入空列 + 表头
    print(f"  Step 4 | 在第 {ins_q_idx} 列插入新列，并命名为 '{header_q}'（整列空白） … / "
          f"Step 4 | Insert new column {ins_q_idx} named '{header_q}' (blank column) ...")
    ws.insert_cols(ins_q_idx, amount=1)
    ws.cell(row=1, column=ins_q_idx).value = header_q
    for r in range(2, max_row + 1):
        ws.cell(row=r, column=ins_q_idx).number_format = "@"

    # 保存到目标
    wb.save(out_path)
    print(f"  💾 已保存清洗结果：{out_path} / Saved cleaned result: {out_path}")
    return out_path


def backup_if_exists(dest_path: str) -> None:
    """若目标已存在，先备份为 .YYYYmmdd_HHMMSS.bak.xlsx"""
    if os.path.exists(dest_path):
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        bak = dest_path.replace(".xlsx", f".{ts}.bak.xlsx")
        try:
            shutil.move(dest_path, bak)
            print(f"  ℹ 发现旧文件，已备份为：{os.path.basename(bak)} / Existing file backed up: {os.path.basename(bak)}")
        except Exception as e:
            print(f"  ⚠ 备份旧文件失败：{e} / Failed to back up existing file: {e}")


def main():
    # 1) 找最新 PR1 Info Record
    latest = find_latest_inforecord(SRC_DIR)
    if not latest:
        print(f"❌ 在 {SRC_DIR} 的最新子目录未找到 PR1 Info Record*.xlsx / "
              f"PR1 Info Record*.xlsx not found in latest subfolder of {SRC_DIR}")
        return

    # 2) 目标准备
    os.makedirs(DEST_DIR, exist_ok=True)
    dest_path = os.path.join(DEST_DIR, DEST_NAME)

    # 3) 等待目标目录空闲（如不需要可直接注释下一段）
    if not wait_dest_clear(DEST_DIR):
        print("⛔ 因目标目录被占用，本次未执行落盘。 / Destination is busy; skipping write.")
        return

    # 4) 备份旧文件并覆盖保存
    backup_if_exists(dest_path)
    process_workbook_and_save(latest, dest_path)

    print("\n🎉 完成： / Completed:")
    print("  源文件：", latest, "/ Source file:", latest)
    print("  目标文件：", dest_path, "/ Destination file:", dest_path)

    tool = SqlAgentTool(server="tcp:10.80.127.71,1433")

    result = tool.run_job(
        job_name="Lumileds BI - SC Purchase Order And Vendor Performance",  # 用完整精确名最稳妥
        archive_dir=r"\\mygbynbyn1msis1\Supply-Chain-Analytics\Data Warehouse\Data Source\SAP\Master Data\Archive",
        timeout=1800,
        poll_interval=3,
        fuzzy=False,  # 若你 later 拿到读 sysjobs 的权限，可改 True
    )
    print(result)


if __name__ == "__main__":
    main()
