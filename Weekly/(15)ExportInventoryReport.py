import os
import sys
import shutil
from typing import List
import pandas as pd
from datetime import date, datetime, timedelta
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

# -------- 配置区 --------
DOWNLOADS = r"C:\Users\70731224\Downloads"
CSV_KEYWORD = "Weekly Trend"      # 模糊匹配关键字，只要包含这个即可

DEST1 = r"\\mygbynbyn1msis2\SCM_Excellence\Weekly Report\Inventory\Tracker"
DEST2 = r"\\mygbynbyn1msis2\LL-PPNC\Weekly inventory"

TEXT_COLS = ["J", "K", "P", "AD", "AN", "BO", "BP", "BS"]  # 强制文本
DATE_COLS = ["L", "M", "S", "T", "AE", "BT"]               # 转成日期 + 格式 dd/mm/yyyy

SUM_COL = "V"                                              # 要求和的列（Excel 列号）

# None = 自动判断月末；True = 强制视为月末；False = 永远不加 _monthend
FORCE_MONTH_END: bool | None = None
# ------------------------


def excel_col_to_index(letter: str) -> int:
    """A -> 1, B -> 2 ..."""
    letter = letter.strip().upper()
    n = 0
    for ch in letter:
        if not ("A" <= ch <= "Z"):
            raise ValueError(f"非法列名: {letter} / Invalid column name: {letter}")
        n = n * 26 + (ord(ch) - 64)
    return n


def ensure_dir(path: str):
    if path and not os.path.exists(path):
        os.makedirs(path, exist_ok=True)


def find_latest_csv(folder: str, keyword: str) -> str | None:
    """在 folder 中按 keyword 模糊匹配最新的 CSV 文件"""
    candidates = []
    for fn in os.listdir(folder):
        if fn.lower().endswith(".csv") and keyword.lower() in fn.lower():
            full = os.path.join(folder, fn)
            candidates.append((full, os.path.getmtime(full)))

    if not candidates:
        return None

    candidates.sort(key=lambda x: x[1], reverse=True)
    newest = candidates[0][0]
    print(f"[INFO] 自动找到最新 CSV：{os.path.basename(newest)} / Latest CSV found: {os.path.basename(newest)}")
    return newest


def read_csv_smart(path: str) -> pd.DataFrame:
    try:
        return pd.read_csv(path, encoding="utf-8-sig")
    except UnicodeDecodeError:
        return pd.read_csv(path, encoding="gbk", errors="ignore")


def is_month_end_day(d: date) -> bool:
    next_month = date(
        d.year + (1 if d.month == 12 else 0),
        1 if d.month == 12 else d.month + 1,
        1,
    )
    return d == (next_month - timedelta(days=1))


def build_filename(today: date) -> str:
    """生成带周号、月末、以及 (Mon'YY Closing) 的文件名"""
    iso_year, iso_week, _ = today.isocalendar()

    # ISO 周号减 1
    week = iso_week - 1
    year = today.year
    if week == 0:
        last_dec31 = date(today.year - 1, 12, 31)
        week = last_dec31.isocalendar()[1]
        year -= 1

    yy = year % 100
    base = f"SAP_Inv_Tracker_Details_ww{week:02d}'{yy:02d}"

    # 月末逻辑
    monthend = FORCE_MONTH_END if FORCE_MONTH_END is not None else is_month_end_day(today)
    if monthend:
        base += "_monthend"

    # 是否是当月第一周的周报
    first_day = date(today.year, today.month, 1)
    first_week_of_month = first_day.isocalendar()[1]
    is_first_week_report = (iso_week == first_week_of_month)

    if is_first_week_report:
        # Closing 月 = 当前月 - 1
        closing_year = today.year
        closing_month = today.month - 1
        if closing_month == 0:
            closing_month = 12
            closing_year -= 1

        month_abbr = [
            "Jan", "Feb", "Mar", "Apr", "May", "Jun",
            "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"
        ][closing_month - 1]

        closing_text = f"({month_abbr}'{str(closing_year)[-2:]} Closing)"
        base += f" {closing_text}"

    return base + ".xlsx"


def convert_date_cols(df: pd.DataFrame,
                      date_cols_letters: List[str],
                      text_cols_letters: List[str]) -> pd.DataFrame:
    """
    把指定列尽量转成 datetime.date（值层面）。

    已改成 dayfirst=True：
    12/01/2025 会按 “12-Jan-2025” 解析，而不是 “01-Dec-2025”。
    """
    n_cols = len(df.columns)
    text_idx_set = {excel_col_to_index(L) - 1 for L in text_cols_letters}

    for L in date_cols_letters:
        idx = excel_col_to_index(L) - 1
        if idx >= n_cols:
            print(f"[警告] 日期列 {L} 超出列数范围，跳过 / Date column {L} out of range, skipped")
            continue
        if idx in text_idx_set:
            print(f"[警告] 列 {L} 在 TEXT_COLS 中，将按文本处理，不转日期 / "
                  f"Column {L} is in TEXT_COLS; keep as text, no date conversion")
            continue

        s = df.iloc[:, idx]
        # 关键修改：dayfirst=True
        dt = pd.to_datetime(s, errors="coerce", dayfirst=True)
        df.iloc[:, idx] = dt.dt.date
        print(f"[INFO] 列 {L} 已在 DataFrame 中转换为日期值 / Column {L} converted to date in DataFrame")

    return df


def write_df_to_xlsx(
    df: pd.DataFrame,
    xlsx_path: str,
    text_cols_letters: List[str],
    date_cols_letters: List[str],
):
    """
    用 openpyxl 写 Excel：
    - 文本列设为文本；
    - 日期列写入真正的日期，并设置为 dd/mm/yyyy 显示。
    """
    wb = Workbook()
    ws: Worksheet = wb.active
    ws.title = "Sheet1"

    nrows, ncols = df.shape

    text_idx_set = {excel_col_to_index(L) - 1 for L in text_cols_letters}
    date_idx_set = {excel_col_to_index(L) - 1 for L in date_cols_letters}

    # 写表头
    for c_idx, name in enumerate(df.columns, start=1):
        ws.cell(row=1, column=c_idx, value=str(name))

    # 写数据
    for r in range(nrows):
        for c in range(ncols):
            val = df.iat[r, c]
            cell = ws.cell(row=r + 2, column=c + 1)

            # 文本列
            if c in text_idx_set:
                cell.number_format = "@"
                cell.value = "" if pd.isna(val) else str(val)
                continue

            # 日期列：写入日期值 + 格式 dd/mm/yyyy
            if c in date_idx_set:
                if pd.isna(val):
                    cell.value = None
                else:
                    v = val
                    if not isinstance(v, (date, datetime)):
                        try:
                            parsed = pd.to_datetime(v, errors="coerce", dayfirst=True)
                            if pd.isna(parsed):
                                v = None
                            else:
                                v = parsed.date()
                        except Exception:
                            v = None
                    cell.value = v
                    if v is not None:
                        # 关键修改：从 "ddmmyyyy" 改成 "dd/mm/yyyy"
                        cell.number_format = "dd/mm/yyyy"
                continue

            # 其他列：原样写入
            cell.value = None if pd.isna(val) else val

    ensure_dir(os.path.dirname(xlsx_path))
    wb.save(xlsx_path)
    print(f"[SAVE] 已保存 Excel：{xlsx_path} / Excel saved: {xlsx_path}")


def main():
    # 1) 找最新 CSV
    csv_path = find_latest_csv(DOWNLOADS, CSV_KEYWORD)
    if not csv_path:
        print("[错误] 未找到匹配 CSV，请检查下载文件夹和关键字。 / "
              "No matching CSV found; check downloads folder and keyword.")
        sys.exit(1)

    print(f"[INFO] 读取 CSV：{csv_path} / Reading CSV: {csv_path}")
    df = read_csv_smart(csv_path)

    # 2) 转换日期列（值）
    df = convert_date_cols(df, DATE_COLS, TEXT_COLS)

    # 3) 打印 SUM_COL 合计
    v_idx = excel_col_to_index(SUM_COL) - 1
    if v_idx >= len(df.columns):
        print(f"[警告] SUM_COL {SUM_COL} 不在文件中，无法求和 / SUM_COL {SUM_COL} not in file; cannot sum")
    else:
        total = pd.to_numeric(df.iloc[:, v_idx], errors="coerce").sum()
        print(f"[SUM] {SUM_COL} 列合计 = {total:,.4f} / {SUM_COL} total = {total:,.4f}")

    input("按回车继续生成 Excel → / Press Enter to continue generating Excel → ")

    today = date.today()
    filename = build_filename(today)

    saveas_path = os.path.join(DEST1, filename)

    # 4) 写 Excel（文本列 + 日期列 dd/mm/yyyy）
    write_df_to_xlsx(df, saveas_path, TEXT_COLS, DATE_COLS)

    # 5) 复制到第二个 shared drive
    ensure_dir(DEST2)
    dst2_path = os.path.join(DEST2, filename)
    shutil.copy2(saveas_path, dst2_path)
    print(f"[COPY] {saveas_path}  ->  {dst2_path}")

    print("✅ 全流程完成！ / Workflow completed!")


if __name__ == "__main__":
    main()
