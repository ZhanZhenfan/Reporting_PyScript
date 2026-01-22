# -*- coding: utf-8 -*-
"""
Step5: 更新 Product_list New
- 在 Winshuttle PR1 reports 找最新 Product_List*.xlsx
- 用上个月 M1M2/Product_list New.xlsx 的 header 替换（严格要求列数一致，保留重复列名）
- 工作表名固定写成 "Product List"
- 保存到 M1M2 文件夹 -> Product_list New.xlsx（不做备份）
"""

import os
from pathlib import Path
import pandas as pd
import winsound
from openpyxl import load_workbook

# 路径
PR1_DIR  = Path(r"\\mygbynbyn1msis1\Supply-Chain-Analytics\Data Warehouse\Data Source\WinShuttle Data Source\PR1 reports")
M1M2_DIR = Path(r"\\mygbynbyn1msis1\Supply-Chain-Analytics\Data Warehouse\Data Source\External\M1M2\Original Raw")
OUT_FILE = M1M2_DIR / "Product_list New.xlsx"
SHEET_NAME = "Product List"  # 固定输出工作表名


def read_excel_header(path, sheet=0):
    """用 openpyxl 直接读第一行表头（保留重复列名）"""
    wb = load_workbook(path, read_only=True)
    ws = wb.worksheets[sheet] if isinstance(sheet, int) else wb[sheet]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    wb.close()
    return headers


def get_latest_productlist(pr1_dir: Path) -> Path:
    files = list(pr1_dir.glob("Product_List*.xlsx"))
    if not files:
        raise FileNotFoundError("PR1 reports 下未找到 Product_List*.xlsx 文件 / No Product_List*.xlsx found in PR1 reports")
    latest = max(files, key=lambda f: f.stat().st_mtime)
    print(f"[STEP 1] 找到最新文件: {latest.name} / Latest file found: {latest.name}")
    return latest


def main():
    print("========== 开始执行 Step5 / Starting Step5 ==========")

    # 1. 找最新源文件
    latest_file = get_latest_productlist(PR1_DIR)

    # 2. 读取旧表头
    print("[STEP 2] 读取旧文件表头 / Read old headers")
    old_headers = read_excel_header(OUT_FILE, SHEET_NAME)
    print(f"         → 旧表头列数: {len(old_headers)} / Old header count: {len(old_headers)}")
    print(f"         → 旧表头前10个: {old_headers[:10]} / First 10 old headers: {old_headers[:10]}")

    # 3. 读取新文件数据（不带表头，跳过第一行）
    print("[STEP 3] 读取新文件数据（跳过首行） / Read new data (skip first row)")
    new_df = pd.read_excel(latest_file, header=None, skiprows=1, dtype=object, engine="openpyxl")
    print(f"         → 新文件列数: {new_df.shape[1]} / New column count: {new_df.shape[1]}")

    # 4. 检查列数一致
    print("[STEP 4] 检查新旧列数是否一致 / Check column count match")
    if new_df.shape[1] != len(old_headers):
        raise ValueError(
            f"列数不一致！新文件={new_df.shape[1]}，旧文件={len(old_headers)} / "
            f"Column count mismatch! New={new_df.shape[1]}, Old={len(old_headers)}"
        )
    print("         → 列数一致 ✅ / Column counts match ✅")

    # 5. 套用旧表头
    print("[STEP 5] 替换表头为旧表头（按位置对齐） / Replace headers with old ones")
    new_df.columns = old_headers
    print("         → 表头替换完成 / Header replacement done")
    print(f"         → 新表头前10个: {list(new_df.columns[:10])} / First 10 new headers: {list(new_df.columns[:10])}")

    # 6. 保存文件
    print("[STEP 6] 保存新文件 / Save new file")
    OUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    tmp = OUT_FILE.with_name(OUT_FILE.stem + ".tmp.xlsx")
    with pd.ExcelWriter(tmp, engine="openpyxl", mode="w") as writer:
        new_df.to_excel(writer, index=False, sheet_name=SHEET_NAME)
    os.replace(tmp, OUT_FILE)
    print(f"         → 已写入: {OUT_FILE} / Written to: {OUT_FILE}")

    print("========== 执行完成 Step5 / Step5 completed ==========")


if __name__ == "__main__":
    main()
    winsound.Beep(1000, 300)
    winsound.Beep(1500, 300)
    winsound.Beep(2000, 300)
