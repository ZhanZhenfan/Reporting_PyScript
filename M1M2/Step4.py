import os
import glob
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from openpyxl.formula.translate import Translator
import winsound


# 1) 路径
ARCHIVE_DIR = r"\\mygbynbyn1msis1\Supply-Chain-Analytics\Data Warehouse\Data Source\SAP\Transactional Data\Outbound\Archive"
TARGET_FILE = r"\\mygbynbyn1msis1\Supply-Chain-Analytics\Data Warehouse\Data Source\External\M1M2\Original Raw\VL06O 2.xlsx"
# TARGET_FILE = r"C:\Users\70731224\OneDrive - Lumileds\Desktop\M1M2\VL06O 2.xlsx"


# === 找最新 VL06O*.xlsx ===
cands = glob.glob(os.path.join(ARCHIVE_DIR, "VL06O*.xlsx"))
if not cands:
    raise FileNotFoundError("Archive 目录没有找到 VL06O*.xlsx / No VL06O*.xlsx found in Archive")
SRC_FILE = max(cands, key=os.path.getmtime)
print("源文件：", SRC_FILE, "/ Source file:", SRC_FILE)
print("目标模板：", TARGET_FILE, "/ Target template:", TARGET_FILE)

# === 打开工作簿（保留格式） ===
src_wb = load_workbook(SRC_FILE, data_only=True)
src_ws = src_wb.active
dst_wb = load_workbook(TARGET_FILE)
dst_ws = dst_wb.active

def last_nonempty_row(ws, col_letter="A", start_row=2):
    col_idx = column_index_from_string(col_letter)
    for r in range(ws.max_row, start_row - 1, -1):
        v = ws.cell(r, col_idx).value
        if v is not None and str(v).strip() != "":
            return r
    return start_row - 1

src_start = 2
src_end = last_nonempty_row(src_ws, "A", src_start)
row_count = max(0, src_end - src_start + 1)
print(f"复制 {row_count} 行。 / Copied {row_count} rows.")

# === 清空目标数据区（保留表头与样式） ===
if dst_ws.max_row > 1:
    dst_ws.delete_rows(2, dst_ws.max_row - 1)

# === 列字母映射（按你确认的“这样才对”） ===
# VL06O 2  ←  VL06O
mapping = {
    "A": ["A"],
    "B": ["B"],
    "C": ["C"],
    "D": ["D"],
    "E": ["P"],
    "F": ["Q"],
    "G": ["AS"],
    "H": ["BQ"],
    "I": ["BR"],
    "J": ["K", "CB"],   # 优先 K，空则用 CB
    "K": ["EF"],
    # L 列稍后写公式
}

def get(ws, r, col_letter):
    return ws.cell(r, column_index_from_string(col_letter)).value

# === 写入映射列（只写值） ===
dst_row = 2
for r in range(src_start, src_end + 1):
    for dcol, scol_list in mapping.items():
        val = None
        for scol in scol_list:
            v = get(src_ws, r, scol)
            if v is not None and str(v).strip() != "":
                val = v
                break
        dst_ws[f"{dcol}{dst_row}"].value = val
    dst_row += 1

# === L 列直接写入公式：L2 =TRIM(A2)，并向下填充 ===
if row_count > 0:
    dst_ws["L2"].value = "=TRIM(A2)"
    # 平移并填充到所有数据行
    for r in range(3, 2 + row_count):
        dst_ws[f"L{r}"].value = Translator("=TRIM(A2)", origin="L2").translate_formula(f"L{r}")

# === 保存 ===
dst_wb.save(TARGET_FILE)
print("完成：数据写入，并已在 L 列填充 =TRIM(A2) 公式。 / Done: data written and L column filled with =TRIM(A2).")



# 频率 (Hz) 和时长 (毫秒)
winsound.Beep(1000, 500)   # 1000Hz，响0.5秒
winsound.Beep(1500, 500)
winsound.Beep(2000, 500)
