# -*- coding: utf-8 -*-
import os
import glob
from datetime import date
from openpyxl import load_workbook

RAW_DIR = r"\\mygbynbyn1msis1\Supply-Chain-Analytics\Data Warehouse\Data Source\External\M1M2\Original Raw"
DST_DIR = r"\\mp1do4ce0373ndz\Customs"

# 每个关键词独立配置：模板名、源数据起始行（含表头行数）
CONFIG = {
    "Scrap": {
        "template": "Scrap",
        "src_data_start_row": 2,
        "dst_header_row": 1,
    },
    "Machinery": {
        "template": "Machinery",
        "src_data_start_row": 6,
        "dst_header_row": 1,
    },
}

def latest_file_contains(folder, keyword):
    pats = [os.path.join(folder, f"*{keyword}*.xlsx"),
            os.path.join(folder, f"*{keyword}*.xlsm")]
    files = []
    for p in pats:
        files.extend(glob.glob(p))
    if not files:
        raise FileNotFoundError(f"找不到包含“{keyword}”的 Excel 文件 / No Excel file containing '{keyword}' found")
    return max(files, key=os.path.getmtime)

def find_template_path(dst_dir, base_name):
    for ext in (".xlsm", ".xlsx"):
        path = os.path.join(dst_dir, base_name + ext)
        if os.path.exists(path):
            return path
    raise FileNotFoundError(f"模板不存在：{base_name}.xlsm/.xlsx @ {dst_dir} / Template not found")

def used_last_row(ws):
    max_r = ws.max_row
    max_c = ws.max_column
    for r in range(max_r, 1, -1):
        for c in range(1, max_c + 1):
            if ws.cell(row=r, column=c).value not in (None, ""):
                return r
    return 1

def header_last_col(ws, header_row=1):
    max_c = ws.max_column
    last_c = 0
    for c in range(1, max_c + 1):
        if ws.cell(row=header_row, column=c).value not in (None, ""):
            last_c = c
    return max(1, last_c)

def prev_month_yyyymm():
    today = date.today()
    y, m = today.year, today.month
    if m == 1:
        y -= 1
        m = 12
    else:
        m -= 1
    return f"{y}{m:02d}"

def copy_values(src_path, dst_path, src_data_start_row, dst_header_row, keyword):
    """
    仅复制值，且根据 keyword 实现列位移/跳列：
      - Machinery: 源从第1列复制，目标从第2列起粘贴
      - Scrap    : 源从第2列复制（跳过源第1列），目标从第2列起粘贴
    复制结束后：对“有数据的行”，把第1列填为 上月YYYYMM。
    保护措施：
      - 不覆盖目标中的公式（以 '=' 开头）
      - 源为空(None/'')则不写，避免清掉模板可能预置的值/验证
    """
    print(f"[COPY] {keyword}: {os.path.basename(src_path)}  →  {os.path.basename(dst_path)} / Copying")

    dst_ext = os.path.splitext(dst_path)[1].lower()
    dst_wb = load_workbook(dst_path, keep_vba=(dst_ext == ".xlsm"), data_only=False)
    # 源仅读数值，避免带入格式/公式
    src_wb = load_workbook(src_path, data_only=True)

    # 表匹配：默认第一张；若同名则优先同名
    dst_ws = dst_wb.worksheets[0]
    src_ws = src_wb.worksheets[0]
    if src_ws.title in dst_wb.sheetnames:
        dst_ws = dst_wb[src_ws.title]

    # 目标有效列按表头确定
    dst_cols = header_last_col(dst_ws, header_row=dst_header_row)

    # 列位移策略（都从目标第2列开始写）
    dst_col_start = 2
    src_col_start = 1 if keyword == "Machinery" else 2

    # 可复制列数
    src_max_cols = src_ws.max_column
    src_cols_available = max(0, src_max_cols - (src_col_start - 1))
    dst_cols_available = max(0, dst_cols - (dst_col_start - 1))
    copy_cols = min(src_cols_available, dst_cols_available)

    # 行范围
    src_last_row = used_last_row(src_ws)
    rows_to_copy = max(0, src_last_row - (src_data_start_row - 1))
    print(
        f"      源起始行: {src_data_start_row}，源末行: {src_last_row} → 行数: {rows_to_copy}；"
        f"列数: {copy_cols}（源从{src_col_start}起 → 目标从{dst_col_start}起） / "
        f"Src start: {src_data_start_row}, src end: {src_last_row} -> rows: {rows_to_copy}; "
        f"cols: {copy_cols} (src from {src_col_start} -> dst from {dst_col_start})"
    )

    # 写入（目标从表头下一行开始），只写“有值且目标不是公式”的格
    dst_row_start = dst_header_row + 1
    if rows_to_copy > 0 and copy_cols > 0:
        for i, r in enumerate(range(src_data_start_row, src_last_row + 1), start=0):
            for j in range(copy_cols):
                src_c = src_col_start + j
                dst_c = dst_col_start + j
                src_val = src_ws.cell(row=r, column=src_c).value
                if src_val in (None, ""):
                    continue  # 不写空，保护模板占位/验证
                dst_cell = dst_ws.cell(row=dst_row_start + i, column=dst_c)
                if isinstance(dst_cell.value, str) and dst_cell.value.startswith("="):
                    continue  # 保护模板公式
                dst_cell.value = src_val  # 仅赋值，不改格式/样式

    # 填充第1列 = 上月YYYYMM（仅对“该行第2列及以后存在任意数据”的行）
    tag = prev_month_yyyymm()
    last_written_row = dst_row_start + max(0, rows_to_copy) - 1
    if last_written_row >= dst_row_start:
        for rr in range(dst_row_start, last_written_row + 1):
            # 行内从第2列起是否有数据（不看公式与空）
            has_data = False
            rightmost_c = dst_col_start + max(0, copy_cols) - 1
            for cc in range(2, max(2, rightmost_c) + 1):
                v = dst_ws.cell(row=rr, column=cc).value
                if v not in (None, ""):
                    has_data = True
                    break
            if has_data:
                # 不改格式，仅写值
                a_cell = dst_ws.cell(row=rr, column=1)
                # 若A列本身是公式也不覆盖
                if not (isinstance(a_cell.value, str) and a_cell.value.startswith("=")):
                    a_cell.value = tag

    dst_wb.save(dst_path)
    src_wb.close()
    dst_wb.close()
    print("      [OK] 写入完成，并已填充第1列 YYYYMM（未改模板格式/公式）。 / "
          "Write complete; filled column 1 with YYYYMM (template formats/formulas unchanged).")

def main():
    for keyword, cfg in CONFIG.items():
        template_name = cfg["template"]
        src_start = cfg["src_data_start_row"]
        dst_header = cfg["dst_header_row"]

        # 1) 找最新原始
        src_file = latest_file_contains(RAW_DIR, keyword)
        print(f"[发现原始] {keyword}: {os.path.basename(src_file)} / Source found")

        # 2) 找模板
        dst_template = find_template_path(DST_DIR, template_name)
        print(f"[匹配模板] {template_name}: {os.path.basename(dst_template)} / Template matched")

        # 3) 复制（含列位移/填A列YYYYMM）
        copy_values(src_file, dst_template, src_start, dst_header, keyword)

    print("✅ 全部完成。 / All done.")

if __name__ == "__main__":
    main()
