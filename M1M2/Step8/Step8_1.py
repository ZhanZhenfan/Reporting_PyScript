# -*- coding: utf-8 -*-
import os
import glob
import shutil
from openpyxl import load_workbook

from M1M2.Step8 import Step8_2, Step8_3

SRC_DIR = r"\\mp1do4ce0373ndz\Customs\Archive"
DST_DIR = r"\\mp1do4ce0373ndz\Customs"

# 只保留 Scrap / Machinery
TARGETS = {
    "Scrap_": "Scrap",
    "Machinery_": "Machinery",
}

def latest_by_prefix(folder, prefix):
    pats = [os.path.join(folder, f"{prefix}*.xlsx"),
            os.path.join(folder, f"{prefix}*.xlsm")]
    files = []
    for p in pats:
        files.extend(glob.glob(p))
    if not files:
        raise FileNotFoundError(f"找不到匹配文件：{prefix}*.xlsx/*.xlsm / No matching files found")
    return max(files, key=os.path.getmtime)

def clear_sheet_values_keep_header(ws, header_row=1):
    """
    清空除表头外的值，保留格式/筛选/合并关系；
    ★ 按你的新要求：从数据区开始清空，但“第1列”一律不清空。
    """
    data_start = header_row + 1
    max_r, max_c = ws.max_row, ws.max_column
    if max_r < data_start:
        return

    # 对合并区域做跳过表，避免清空非左上角单元格
    skip_coords = set()
    for mr in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = mr.bounds
        if max_row >= data_start:
            for r in range(max(min_row, data_start), max_row + 1):
                for c in range(min_col, max_col + 1):
                    if not (r == min_row and c == min_col):
                        skip_coords.add((r, c))

    for r in range(data_start, max_r + 1):
        for c in range(1, max_c + 1):
            # ★ 不清空第1列
            if c == 1:
                continue
            if (r, c) in skip_coords:
                continue
            ws.cell(row=r, column=c).value = None

def process_one(src_path, base_name):
    ext = os.path.splitext(src_path)[1].lower()
    if ext not in (".xlsx", ".xlsm"):
        print(f"[跳过] 不支持的格式：{src_path} / Skipped: unsupported format")
        return
    dst_path = os.path.join(DST_DIR, base_name + ext)

    shutil.copy2(src_path, dst_path)

    wb = load_workbook(dst_path, keep_vba=(ext == ".xlsm"), data_only=False)
    for ws in wb.worksheets:
        clear_sheet_values_keep_header(ws, header_row=1)
    wb.save(dst_path)
    print(f"[OK] {os.path.basename(src_path)} → {dst_path}")

def main():
    os.makedirs(DST_DIR, exist_ok=True)
    for prefix, base in TARGETS.items():
        src = latest_by_prefix(SRC_DIR, prefix)
        print(f"[处理] {os.path.basename(src)} / Processing")
        process_one(src, base)

    try:
        import winsound
        winsound.Beep(1000, 250); winsound.Beep(1500, 250); winsound.Beep(2000, 350)
    except Exception:
        print("\a")
    print("✅ 全部完成：", DST_DIR, "/ All done:", DST_DIR)

if __name__ == "__main__":
    main()
    Step8_2.main()
    Step8_3.main()
