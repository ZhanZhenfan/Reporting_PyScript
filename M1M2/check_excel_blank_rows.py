from __future__ import annotations
from pathlib import Path
import re
import shutil
from datetime import datetime
import sys
import time
import traceback

import openpyxl
import pythoncom  # type: ignore
from win32com.client import Dispatch, gencache  # type: ignore

# ===== Configuration =====
BASE_DIR = Path(r"\\mygbynbyn1msis1\Supply-Chain-Analytics\Data Warehouse\Data Source\External\M1M2\Original Raw")
BASE_NAMES = [
    "M2 ZSD VL06O",
    "ZVF05",
    "VL06i",
    "MB51-M1",
    "MB51-M2",
    "MM60",
]
STRICT_MUST_FIND_ALL = True  # Exit with error if any base name not found

# ===== Logging Tools =====
LOG_DIR = BASE_DIR / "_logs"
LOG_DIR.mkdir(parents=True, exist_ok=True)
LOG_PATH = LOG_DIR / f"repair_m1m2_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"

def _now():
    return datetime.now().strftime("%H:%M:%S")

def log(msg: str):
    line = f"[{_now()}] {msg}"
    print(line, flush=True)
    try:
        with open(LOG_PATH, "a", encoding="utf-8") as f:
            f.write(line + "\n")
    except Exception:
        pass

class Timer:
    def __init__(self, label: str):
        self.label = label
        self.t0 = None
    def __enter__(self):
        self.t0 = time.time()
        return self
    def __exit__(self, exc_type, exc, tb):
        if self.t0 is not None:
            dt = time.time() - self.t0
            log(f"⏱  {self.label} took {dt:.2f}s")

# ===== Utilities =====
_WS_PATTERN = re.compile(r"^[\u0009\u000A\u000D\u0020\u00A0\u1680\u180E\u2000-\u200B\u202F\u205F\u3000\uFEFF]*$")
LEADING_ZERO = re.compile(r"^0\d+$")

def _has_value(v) -> bool:
    if v is None: return False
    if isinstance(v, str): return not bool(_WS_PATTERN.match(v))
    return True

def true_last_row_col(ws):
    max_r, max_c = ws.max_row or 0, ws.max_column or 0
    if max_r == 0 or max_c == 0: return 0, 0
    last_r = 0
    for r in range(max_r, 0, -1):
        if any(_has_value(ws.cell(r, c).value) for c in range(1, max_c+1)):
            last_r = r; break
    if last_r == 0: return 0, 0
    last_c = 0
    for c in range(max_c, 0, -1):
        if any(_has_value(ws.cell(r, c).value) for r in range(1, max_r+1)):
            last_c = c; break
    return last_r, last_c

def safe_lastcell(ws):
    try:
        lc = ws.Cells.SpecialCells(11)  # xlCellTypeLastCell
        return int(lc.Row), int(lc.Column)
    except Exception:
        try:
            ur = ws.UsedRange
            r = int(ur.Row) + int(ur.Rows.Count) - 1
            c = int(ur.Column) + int(ur.Columns.Count) - 1
            if r >= 1 and c >= 1: return r, c
        except Exception:
            pass
        xlUp, xlToLeft = -4162, -4159
        last_row = ws.Cells(ws.Rows.Count, 1).End(xlUp).Row
        last_col = ws.Cells(1, ws.Columns.Count).End(xlToLeft).Column
        return int(last_row), int(last_col)

def _flatten_1col_vals(vals, rows):
    if rows == 1:
        return [vals]
    if isinstance(vals, tuple) and vals and isinstance(vals[0], tuple):
        return [row[0] if isinstance(row, tuple) and row else None for row in vals]
    return list(vals) if isinstance(vals, (list, tuple)) else [vals]

def column_should_be_text(src_ws, rows, col_idx):
    fmt = str(src_ws.Columns(col_idx).NumberFormat)
    if fmt.strip() == "@":
        return True
    saw_num = saw_non_num = saw_text = saw_leading_zero = False
    vals = src_ws.Range(src_ws.Cells(1, col_idx), src_ws.Cells(rows, col_idx)).Value
    vals = _flatten_1col_vals(vals, rows)
    for v in vals:
        if v is None or v == "":
            continue
        if isinstance(v, str):
            saw_text = True
            if LEADING_ZERO.match(v): saw_leading_zero = True
            try: float(v); saw_num = True
            except Exception: saw_non_num = True
        else:
            try: float(v); saw_num = True
            except Exception: saw_non_num = True
    if saw_text or saw_leading_zero: return True
    if saw_num and saw_non_num: return True
    return False

def copy_preserve_types(src_ws, dst_ws, rows, cols):
    if rows == 0 or cols == 0:
        return
    milestones = {max(1, int(cols*0.25)), max(1, int(cols*0.5)), max(1, int(cols*0.75)), cols}
    for c in range(1, cols+1):
        dst_col = dst_ws.Columns(c)
        src_col = src_ws.Columns(c)
        as_text = column_should_be_text(src_ws, rows, c)
        dst_col.NumberFormat = "@" if as_text else src_col.NumberFormat
        src_rng = src_ws.Range(src_ws.Cells(1, c), src_ws.Cells(rows, c))
        dst_rng = dst_ws.Range(dst_ws.Cells(1, c), dst_ws.Cells(rows, c))
        dst_rng.Value = src_rng.Value
        dst_col.ColumnWidth = src_col.ColumnWidth
        if c in milestones:
            pct = int(c/cols*100)
            log(f"      · Column copy progress {c}/{cols} ({pct}%)")

def rebuild_sheet_preserve_types(wb_com, ws_name, true_r, true_c):
    src = wb_com.Worksheets(ws_name)
    tmp = wb_com.Worksheets.Add(After=src)
    tmp.Name = f"{ws_name}__tmp"
    if true_r > 0 and true_c > 0:
        with Timer(f"Sheet {ws_name} copy {true_r}x{true_c}"):
            copy_preserve_types(src, tmp, true_r, true_c)
    src.Delete()
    tmp.Name = ws_name

def compute_true_regions(file_path: Path) -> dict[str, tuple[int, int]]:
    with Timer("openpyxl compute true regions"):
        wb_py = openpyxl.load_workbook(str(file_path), data_only=True, read_only=False)
        out = {}
        for ws in wb_py.worksheets:
            r, c = true_last_row_col(ws)
            log(f"    · [{ws.title}] TrueRange = {r}x{c}")
            out[ws.title] = (r, c)
        return out

def fuzzy_pick_latest(base_dir: Path, base_name: str) -> Path | None:
    pat = re.compile(rf"^{re.escape(base_name)}([\s_-].*)?\.xlsx$", re.IGNORECASE)
    candidates = []
    for p in base_dir.glob("*.xlsx"):
        if p.name.startswith("~$"):  # Exclude temporary files
            continue
        if pat.match(p.name):
            candidates.append(p)
    if not candidates:
        return None
    candidates.sort(key=lambda x: x.stat().st_mtime, reverse=True)
    return candidates[0]

def ensure_backup_dir(base_dir: Path) -> Path:
    bk = base_dir / f"_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    bk.mkdir(exist_ok=True)
    return bk

def repair_file_inplace(file_path: Path, app) -> None:
    log(f"— Opening: {file_path.name}")
    sheets_true = compute_true_regions(file_path)
    wb = app.Workbooks.Open(str(file_path))
    names = [ws.Name for ws in wb.Worksheets]
    log(f"— Rebuilding: {file_path.name} (total {len(names)} sheets)")
    for name in names:
        r, c = sheets_true.get(name, (0, 0))
        if r == 0 and c == 0:
            if wb.Worksheets.Count > 1:
                wb.Worksheets(name).Delete()
                log(f"  [Delete empty sheet] {name}")
            else:
                log(f"  [Keep empty sheet] {name} (only one sheet left)")
            continue
        log(f"  [Start] Sheet {name} - TrueRange {r}x{c}")
        rebuild_sheet_preserve_types(wb, name, r, c)
        log(f"  [Complete] Sheet {name}")
    wb.Close(SaveChanges=True)
    log(f"— Save completed: {file_path.name}")

def verify_lastcell(file_path: Path, app) -> None:
    wb = app.Workbooks.Open(str(file_path))
    log("— Verifying Ctrl+End:")
    for ws in wb.Worksheets:
        lr, lc = safe_lastcell(ws)
        log(f"    [{ws.Name}] LastCell({lr},{lc})")
    wb.Close(SaveChanges=False)

# ===== Main Process =====
def main():
    if not BASE_DIR.exists():
        log(f"Directory does not exist: {BASE_DIR}")
        sys.exit(1)

    log("Phase 1/4: Scanning for latest matching files")
    selected: list[tuple[str, Path]] = []
    missing: list[str] = []

    for base in BASE_NAMES:
        p = fuzzy_pick_latest(BASE_DIR, base)
        if p:
            t = datetime.fromtimestamp(p.stat().st_mtime).strftime("%Y-%m-%d %H:%M:%S")
            log(f"  ✓ {base} -> {p.name}  [modified {t}]")
            selected.append((base, p))
        else:
            log(f"  ✗ Not found: {base}*.xlsx")
            missing.append(base)

    if STRICT_MUST_FIND_ALL and missing:
        log(f"❌ The following files were not found (total {len(missing)}/{len(BASE_NAMES)}):")
        for b in missing:
            log(f"   - {b}*.xlsx")
        log("Program terminated. Please confirm files are placed in the directory before running again.")
        sys.exit(2)

    if not selected:
        log("No target files found, exiting.")
        sys.exit(2)

    log("Phase 2/4: Creating unified backup")
    backup_dir = ensure_backup_dir(BASE_DIR)
    log(f"  Backup directory: {backup_dir}")
    for _, p in selected:
        with Timer(f"Backup {p.name}"):
            shutil.copy2(p, backup_dir / p.name)
        log(f"  Backed up: {p.name}")

    log("Phase 3/4: Excel COM rebuild (preserving types)")
    pythoncom.CoInitialize()
    app = None
    try:
        gencache.EnsureDispatch("Excel.Application")
        app = Dispatch("Excel.Application")
        app.Visible = False
        app.DisplayAlerts = False
        for idx, (_, p) in enumerate(selected, 1):
            log(f"\n=== File {idx}/{len(selected)}: {p.name} ===")
            with Timer(f"{p.name} total rebuild time"):
                try:
                    repair_file_inplace(p, app)
                except Exception as e:
                    log(f"!!! Processing failed: {p.name} | {e}")
                    log(traceback.format_exc())
    finally:
        try:
            if app: app.Quit()
        except Exception:
            pass
        pythoncom.CoUninitialize()

    log("\nPhase 4/4: Result verification")
    pythoncom.CoInitialize()
    app2 = None
    try:
        app2 = Dispatch("Excel.Application")
        app2.Visible = False
        app2.DisplayAlerts = False
        for _, p in selected:
            with Timer(f"{p.name} verification"):
                verify_lastcell(p, app2)
    finally:
        try:
            if app2: app2.Quit()
        except Exception:
            pass
        pythoncom.CoUninitialize()

    log("\n✅ All tasks completed. Detailed log available at: {}".format(LOG_PATH))

if __name__ == "__main__":
    main